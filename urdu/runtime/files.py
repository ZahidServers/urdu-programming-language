"""
File utilities for the Urdu Programming Language.
Provides Urdu-named wrappers for zipfile (stdlib) and openpyxl (Excel).

Usage:
    درآمد { زپ_فائل, ایکسل_کتاب, زپ_بناؤ, ایکسل_پڑھو } سے "اردو/فائلیں";
"""

from __future__ import annotations
import os
import shutil
import zipfile as _zf
from pathlib import Path as _Path
from typing import Optional, List, Union, Callable
from .builtins import _UrduObj


# ══════════════════════════════════════════════════════════════════════════════
#  ZIP
# ══════════════════════════════════════════════════════════════════════════════

class زپ_فائل:
    """
    zipfile wrapper.

    طریقہ: "r" پڑھنا  |  "w" لکھنا  |  "a" شامل کرنا
    """

    def __init__(self, راستہ: str, طریقہ: str = "r"):
        self._path = راستہ
        self._mode = طریقہ
        self._zf   = _zf.ZipFile(راستہ, طریقہ)

    # ── Listing ───────────────────────────────────────────────────────────────

    def فہرست(self) -> List[str]:
        """List all file names inside the archive."""
        return self._zf.namelist()

    def معلومات(self, نام: str) -> _UrduObj:
        """Get metadata for one entry (size, compress_size, filename)."""
        info = self._zf.getinfo(نام)
        return _UrduObj({
            "نام":            info.filename,
            "اصل_سائز":       info.file_size,
            "سکڑا_سائز":      info.compress_size,
            "تاریخ":          str(info.date_time),
        })

    # ── Reading ───────────────────────────────────────────────────────────────

    def پڑھو(self, نام: str) -> bytes:
        """Read a file from the archive as bytes."""
        return self._zf.read(نام)

    def متن_پڑھو(self, نام: str, انکوڈنگ: str = "utf-8") -> str:
        """Read a file from the archive as text."""
        return self._zf.read(نام).decode(انکوڈنگ)

    # ── Extraction ────────────────────────────────────────────────────────────

    def نکالو(self, نام: str, منزل: str = ".") -> str:
        """Extract one file. Returns the extracted path."""
        return self._zf.extract(نام, منزل)

    def سب_نکالو(self, منزل: str = ".") -> List[str]:
        """Extract all files. Returns list of extracted paths."""
        self._zf.extractall(منزل)
        return [os.path.join(منزل, n) for n in self._zf.namelist()]

    # ── Writing ───────────────────────────────────────────────────────────────

    def شامل(self, فائل_راستہ: str, زپ_نام: str = None):
        """Add an existing file to the archive."""
        self._zf.write(فائل_راستہ, arcname=زپ_نام or os.path.basename(فائل_راستہ))

    def متن_شامل(self, زپ_نام: str, مواد: str, انکوڈنگ: str = "utf-8"):
        """Add a string as a file inside the archive."""
        self._zf.writestr(زپ_نام, مواد.encode(انکوڈنگ))

    def بائٹس_شامل(self, زپ_نام: str, مواد: bytes):
        """Add raw bytes as a file inside the archive."""
        self._zf.writestr(زپ_نام, مواد)

    # ── Close ─────────────────────────────────────────────────────────────────

    def بند(self):
        """Close the archive."""
        self._zf.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.بند()

    def __repr__(self):
        return f"<زپ_فائل '{self._path}' [{len(self.فہرست())} فائلیں]>"


# ── Standalone ZIP helpers ────────────────────────────────────────────────────

def زپ_بناؤ(زپ_راستہ: str, فائلیں: list) -> str:
    """
    Create a zip archive from a list of file paths (or (path, arcname) tuples).
    Returns the zip path.
    """
    with _zf.ZipFile(زپ_راستہ, "w", _zf.ZIP_DEFLATED) as zf:
        for item in فائلیں:
            if isinstance(item, (list, tuple)):
                src, arcname = item[0], item[1]
            else:
                src, arcname = item, os.path.basename(item)
            zf.write(src, arcname=arcname)
    return زپ_راستہ


def زپ_نکالو(زپ_راستہ: str, منزل: str = ".") -> List[str]:
    """Extract all files from a zip archive. Returns extracted paths."""
    with _zf.ZipFile(زپ_راستہ, "r") as zf:
        zf.extractall(منزل)
        return [os.path.join(منزل, n) for n in zf.namelist()]


def زپ_فہرست(زپ_راستہ: str) -> List[str]:
    """List file names inside a zip without extracting."""
    with _zf.ZipFile(زپ_راستہ, "r") as zf:
        return zf.namelist()


# ══════════════════════════════════════════════════════════════════════════════
#  Excel (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError("openpyxl نصب کریں:  urdu نصب اردو/فائلیں")


class ایکسل_ورق:
    """Wraps an openpyxl Worksheet with Urdu API."""

    def __init__(self, ws):
        self._ws = ws

    @property
    def نام(self) -> str:
        return self._ws.title

    @property
    def صف_تعداد(self) -> int:
        return self._ws.max_row

    @property
    def کالم_تعداد(self) -> int:
        return self._ws.max_column

    def خانہ(self, صف: int, کالم: int):
        """Get cell value at (row, col) — 1-indexed."""
        return self._ws.cell(row=صف, column=کالم).value

    def خانہ_مقرر(self, صف: int, کالم: int, قدر):
        """Set cell value at (row, col) — 1-indexed."""
        self._ws.cell(row=صف, column=کالم).value = قدر

    def قطار_پڑھو(self, صف: int) -> List:
        """Return all values in a row as a list."""
        return [self._ws.cell(row=صف, column=c).value
                for c in range(1, self._ws.max_column + 1)]

    def کالم_پڑھو(self, کالم: int) -> List:
        """Return all values in a column as a list."""
        return [self._ws.cell(row=r, column=کالم).value
                for r in range(1, self._ws.max_row + 1)]

    def سب_ڈیٹا(self) -> List[List]:
        """Return all rows as list of lists (includes header)."""
        return [[cell.value for cell in row] for row in self._ws.iter_rows()]

    def قطار_شامل(self, قدریں: list):
        """Append a row of values."""
        self._ws.append(قدریں)

    def __repr__(self):
        return f"<ایکسل_ورق '{self.نام}' {self.صف_تعداد}x{self.کالم_تعداد}>"


class ایکسل_کتاب:
    """
    openpyxl Workbook wrapper.

    ایکسل_کتاب()           — نئی کتاب
    ایکسل_کتاب("file.xlsx") — موجودہ فائل کھولو
    """

    def __init__(self, راستہ: str = None):
        ox = _openpyxl()
        if راستہ:
            self._wb   = ox.load_workbook(راستہ)
            self._path = راستہ
        else:
            self._wb   = ox.Workbook()
            self._path = None

    # ── Sheets ────────────────────────────────────────────────────────────────

    def ورق_فہرست(self) -> List[str]:
        """List all sheet names."""
        return self._wb.sheetnames

    def ورق(self, نام: str = None) -> ایکسل_ورق:
        """Get sheet by name. If None, returns the active sheet."""
        ws = self._wb[نام] if نام else self._wb.active
        return ایکسل_ورق(ws)

    def ورق_بنائیں(self, نام: str) -> ایکسل_ورق:
        """Create and return a new sheet."""
        ws = self._wb.create_sheet(نام)
        return ایکسل_ورق(ws)

    def ورق_حذف(self, نام: str):
        """Delete a sheet by name."""
        del self._wb[نام]

    # ── Save / Close ──────────────────────────────────────────────────────────

    def محفوظ(self, راستہ: str = None) -> str:
        """Save workbook. Uses original path if none given."""
        dest = راستہ or self._path
        if not dest:
            raise ValueError("محفوظ کرنے کے لیے راستہ دیں")
        self._wb.save(dest)
        self._path = dest
        return dest

    def بند(self):
        """Close the workbook."""
        self._wb.close()

    def __repr__(self):
        sheets = ", ".join(self.ورق_فہرست())
        return f"<ایکسل_کتاب ورق=[{sheets}]>"


# ── Standalone Excel helpers ──────────────────────────────────────────────────

def ایکسل_پڑھو(راستہ: str, ورق_نام: str = None) -> List[List]:
    """
    Read an Excel file and return all rows as a list of lists.
    ورق_نام=None uses the active sheet.
    """
    ox = _openpyxl()
    wb = ox.load_workbook(راستہ, read_only=True, data_only=True)
    ws = wb[ورق_نام] if ورق_نام else wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()
    return data


def ایکسل_لکھو(ڈیٹا: list, راستہ: str, ورق_نام: str = "Sheet1") -> str:
    """
    Write a list-of-lists (or list-of-dicts) to an Excel file.
    Returns the saved path.
    """
    ox = _openpyxl()
    wb = ox.Workbook()
    ws = wb.active
    ws.title = ورق_نام
    for قطار in ڈیٹا:
        if isinstance(قطار, dict):
            ws.append(list(قطار.values()))
        else:
            ws.append(list(قطار))
    wb.save(راستہ)
    wb.close()
    return راستہ


# ══════════════════════════════════════════════════════════════════════════════
#  File Manager  (pathlib + shutil wrappers)
# ══════════════════════════════════════════════════════════════════════════════

# ── Path operations ───────────────────────────────────────────────────────────

def راستہ_ملائیں(*اجزاء: str) -> str:
    """Join path segments. Equivalent to os.path.join / Path(...)."""
    return str(_Path(*اجزاء))


def مطلق_راستہ(راستہ: str) -> str:
    """Return the absolute path."""
    return str(_Path(راستہ).resolve())


def والد_فولڈر(راستہ: str) -> str:
    """Return the parent directory of a path."""
    return str(_Path(راستہ).parent)


def فائل_نام(راستہ: str) -> str:
    """Return the filename (last component) of a path."""
    return _Path(راستہ).name


def فائل_لاحقہ(راستہ: str) -> str:
    """Return the file extension, e.g. '.txt'."""
    return _Path(راستہ).suffix


def فائل_موجود(راستہ: str) -> bool:
    """True if path exists (file or directory)."""
    return _Path(راستہ).exists()


def فائل_ہے(راستہ: str) -> bool:
    """True if path is a regular file."""
    return _Path(راستہ).is_file()


def فولڈر_ہے(راستہ: str) -> bool:
    """True if path is a directory."""
    return _Path(راستہ).is_dir()


def فائل_سائز(راستہ: str) -> int:
    """Return file size in bytes."""
    return _Path(راستہ).stat().st_size


# ── File operations ───────────────────────────────────────────────────────────

def فائل_پڑھو(راستہ: str, انکوڈنگ: str = "utf-8") -> str:
    """Read a text file and return its contents."""
    return _Path(راستہ).read_text(encoding=انکوڈنگ)


def فائل_لکھو(راستہ: str, مواد: str, انکوڈنگ: str = "utf-8") -> str:
    """Write text to a file (overwrites). Returns the path."""
    _Path(راستہ).write_text(مواد, encoding=انکوڈنگ)
    return راستہ


def فائل_شامل(راستہ: str, مواد: str, انکوڈنگ: str = "utf-8") -> str:
    """Append text to a file. Returns the path."""
    with open(راستہ, "a", encoding=انکوڈنگ) as fh:
        fh.write(مواد)
    return راستہ


def فائل_کاپی(ماخذ: str, منزل: str) -> str:
    """Copy a file from source to destination. Returns destination path."""
    return str(shutil.copy2(ماخذ, منزل))


def فائل_منتقل(ماخذ: str, منزل: str) -> str:
    """Move a file or directory. Returns destination path."""
    return str(shutil.move(ماخذ, منزل))


def فائل_حذف(راستہ: str) -> None:
    """Delete a file."""
    _Path(راستہ).unlink(missing_ok=True)


def فائل_نام_بدلیں(پرانا: str, نیا: str) -> str:
    """Rename a file or directory. Returns new path."""
    return str(_Path(پرانا).rename(نیا))


# ── Directory operations ──────────────────────────────────────────────────────

def فولڈر_بنائیں(راستہ: str, *, درمیانی: bool = True) -> str:
    """Create directory (and parents by default). Returns the path."""
    _Path(راستہ).mkdir(parents=درمیانی, exist_ok=True)
    return راستہ


def فولڈر_کاپی(ماخذ: str, منزل: str) -> str:
    """Copy an entire directory tree. Returns destination path."""
    return str(shutil.copytree(ماخذ, منزل))


def فولڈر_حذف(راستہ: str) -> None:
    """Delete a directory and all its contents."""
    shutil.rmtree(راستہ, ignore_errors=True)


def فولڈر_خالی_کریں(راستہ: str) -> None:
    """Delete all contents of a directory but keep the directory itself."""
    p = _Path(راستہ)
    for child in p.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


# ── Search / listing ──────────────────────────────────────────────────────────

def فولڈر_فہرست(راستہ: str = ".") -> List[str]:
    """List all entries in a directory (names only)."""
    return [e.name for e in _Path(راستہ).iterdir()]


def فولڈر_فائلیں(راستہ: str = ".") -> List[str]:
    """List only files in a directory (names only)."""
    return [e.name for e in _Path(راستہ).iterdir() if e.is_file()]


def فائل_تلاش(فولڈر: str, نمونہ: str = "*") -> List[str]:
    """
    Recursively search for files matching a glob pattern.
    Example: فائل_تلاش(".", "*.txt") → ["./notes.txt", ...]
    """
    return [str(p) for p in _Path(فولڈر).rglob(نمونہ)]


def فائل_تلاش_متن(فولڈر: str, متن: str, نمونہ: str = "*.txt",
                   انکوڈنگ: str = "utf-8") -> List[dict]:
    """
    Search for text inside files matching a glob pattern.
    Returns list of {فائل, سطر, مواد} dicts for each matching line.
    """
    results: List[dict] = []
    for path in _Path(فولڈر).rglob(نمونہ):
        if not path.is_file():
            continue
        try:
            for i, line in enumerate(path.read_text(encoding=انکوڈنگ).splitlines(), 1):
                if متن in line:
                    results.append({"فائل": str(path), "سطر": i, "مواد": line.strip()})
        except (OSError, UnicodeDecodeError):
            pass
    return results


# ── File watcher (optional watchdog) ─────────────────────────────────────────

class فائل_نگران:
    """
    Watch a directory for changes. Requires the `watchdog` package.

    Usage:
        ن = فائل_نگران(".", تبدیلی=lambda e: پرنٹ(e))
        ن.شروع()
        ...
        ن.روکو()
    """

    def __init__(self, فولڈر: str = ".",
                 تبدیلی: Callable = None,
                 بنائی: Callable = None,
                 حذف: Callable = None,
                 منتقل: Callable = None,
                 بار_بار: bool = True):
        try:
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler
        except ImportError:
            raise ImportError("فائل نگران کے لیے چلائیں: pip install watchdog")

        self._path = str(_Path(فولڈر).resolve())
        self._recursive = بار_بار

        class _Handler(FileSystemEventHandler):
            def on_modified(self, event):
                if تبدیلی:
                    تبدیلی(event.src_path)

            def on_created(self, event):
                if بنائی:
                    بنائی(event.src_path)

            def on_deleted(self, event):
                if حذف:
                    حذف(event.src_path)

            def on_moved(self, event):
                if منتقل:
                    منتقل(event.src_path, event.dest_path)

        self._observer = Observer()
        self._observer.schedule(_Handler(), self._path, recursive=self._recursive)

    def شروع(self) -> "فائل_نگران":
        self._observer.start()
        return self

    def روکو(self) -> None:
        self._observer.stop()
        self._observer.join()

    def __enter__(self):
        return self.شروع()

    def __exit__(self, *_):
        self.روکو()


# ── Exports ───────────────────────────────────────────────────────────────────

__all__ = [
    # ZIP
    "زپ_فائل", "زپ_بناؤ", "زپ_نکالو", "زپ_فہرست",
    # Excel
    "ایکسل_کتاب", "ایکسل_ورق", "ایکسل_پڑھو", "ایکسل_لکھو",
    # File Manager — path ops
    "راستہ_ملائیں", "مطلق_راستہ", "والد_فولڈر", "فائل_نام", "فائل_لاحقہ",
    "فائل_موجود", "فائل_ہے", "فولڈر_ہے", "فائل_سائز",
    # File Manager — file ops
    "فائل_پڑھو", "فائل_لکھو", "فائل_شامل",
    "فائل_کاپی", "فائل_منتقل", "فائل_حذف", "فائل_نام_بدلیں",
    # File Manager — directory ops
    "فولڈر_بنائیں", "فولڈر_کاپی", "فولڈر_حذف", "فولڈر_خالی_کریں",
    # File Manager — search/listing
    "فولڈر_فہرست", "فولڈر_فائلیں", "فائل_تلاش", "فائل_تلاش_متن",
    # File watcher
    "فائل_نگران",
    # English aliases
    "ZipFile", "ExcelBook", "ExcelSheet",
    "zip_create", "zip_extract", "zip_list",
    "excel_read", "excel_write",
    "path_join", "abs_path", "parent_dir", "file_name", "file_ext",
    "file_exists", "is_file", "is_dir", "file_size",
    "file_read", "file_write", "file_append",
    "file_copy", "file_move", "file_delete", "file_rename",
    "dir_create", "dir_copy", "dir_delete", "dir_clear",
    "dir_list", "dir_files", "file_find", "file_find_text",
    "FileWatcher",
]

# English aliases
ZipFile    = زپ_فائل
ExcelBook  = ایکسل_کتاب
ExcelSheet = ایکسل_ورق
zip_create  = زپ_بناؤ
zip_extract = زپ_نکالو
zip_list    = زپ_فہرست
excel_read  = ایکسل_پڑھو
excel_write = ایکسل_لکھو

path_join    = راستہ_ملائیں
abs_path     = مطلق_راستہ
parent_dir   = والد_فولڈر
file_name    = فائل_نام
file_ext     = فائل_لاحقہ
file_exists  = فائل_موجود
is_file      = فائل_ہے
is_dir       = فولڈر_ہے
file_size    = فائل_سائز
file_read    = فائل_پڑھو
file_write   = فائل_لکھو
file_append  = فائل_شامل
file_copy    = فائل_کاپی
file_move    = فائل_منتقل
file_delete  = فائل_حذف
file_rename  = فائل_نام_بدلیں
dir_create   = فولڈر_بنائیں
dir_copy     = فولڈر_کاپی
dir_delete   = فولڈر_حذف
dir_clear    = فولڈر_خالی_کریں
dir_list     = فولڈر_فہرست
dir_files    = فولڈر_فائلیں
file_find    = فائل_تلاش
file_find_text = فائل_تلاش_متن
FileWatcher  = فائل_نگران
